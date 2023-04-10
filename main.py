# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


from openpyxl import *
from teamClass import Team
from openpyxl import workbook

path = 'C:\\Users\\dakot\\Desktop\\BigBoard.xlsx'
wb = load_workbook(filename = path)
ws = wb['Sheet1']
drafted_list = []

def main():
    i = 1
    for team in r1order:
        print(str(i) + ": " + team.name + ": " + getPick(team))
        i += 1
    for team in r2order:
        print(str(i) + ": " + team.name + ": " + getPick(team))
        i += 1
# afc teams



def getPick(team):
    for cell in ws['B']:
        tempPick = ws.cell(row = cell.row, column=3).value
        if cell.value in team.needs:
            if tempPick not in drafted_list:
                draftPick = tempPick
                drafted_list.append(draftPick)
                team.needs.remove(cell.value)
                #print(draftPick)
                return draftPick

                break



#north
pit = Team("Steelers", ["CB","OT", "LB", "DL"])
bal = Team("Ravens", ["WR", "CB", "DL",])
cle = Team("Browns", ["WR", "DL", "S"])
cin = Team("Bengals", ["TE", "RT", "CB", "S"])

#west
kc = Team("Chiefs", ["OT", "EDGE", "WR", "DL"])
lac = Team("Chargers", ["DL", "WR", "EDGE"])
den = Team("Broncos", ["LB", "WR"])
lv = Team("Raiders", ["QB", "iOL", "CB"])

#south
ten = Team("Titans", ["OT", "EDGE", "WR"])
ind = Team("Colts", ["QB", "iOL", "OT"])
jax = Team("Jaguars", ["CB", "OT", "RB"])
hou = Team("Texans", ["QB", "WR", "DL"])

#east
buf = Team("Bills", ["iOL", "LB", "RB"])
mia = Team("Dolphins", ["S", "LB", "DL"])
nyj = Team("Jets", ["OT", "DL", "LB"])
ne = Team("Patriots", ["WR", "OT", "CB"])

# nfc teams

#north
min = Team("Vikings", ["CB", "LB", "WR"])
chi = Team("Bears", ["DL", "EDGE", "OT",])
gb = Team("Packers", ["TE", "WR", "S"])
det = Team("Lions", ["DL", "S", "WR", "CB"])

#west
sea = Team("Seahawks", ["EDGE", "iOL", "DL", "WR"])
lar = Team("Rams", ["WR", "RB", "TE"])
sf = Team("49ers", ["CB", "LB", "iOL"])
ari = Team("Cardinals", ["EDGE", "DL", "CB"])

#south
atl = Team("Falcons", ["EDGE", "WR", "DL"])
car = Team("Panthers", ["QB", "WR"])
nos = Team("Saints", ["EDGE", "DL"])
tb = Team("Buccaneers", ["OT", "QB", "CB"])

#east
phi = Team("Eagles", ["EDGE", "CB", "iOL"])
dal = Team("Cowboys", ["DL", "RB", "WR"])
was = Team("Commanders", ["QB", "OT", "CB"])
nyg = Team("Giants", ["CB", "WR", "S"])

#draft order
r1order = [car, hou, ari, ind, sea, det, lv, atl, chi, phi, ten, hou, nyj, ne, gb, was, pit, det, tb, sea, lac, bal, min, jax, nyg, dal, buf, cin, nos, phi, kc]
r2order = [pit, hou, ari, ind, lar, sea, lv, car, nos, ten, nyj, nyj, atl, gb, ne, was, det, pit, tb, mia, sea, chi, lac, det, jax, nyg, dal, buf, cin, chi, phi, kc]
#r3order = []
#r4order = []
#r5order = []
#r6order = []
#r7order = []

if __name__ == '__main__':
    main()
